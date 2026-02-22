from rest_framework import viewsets, filters, permissions
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework import status
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.permissions import IsAdminUser, IsAuthenticatedOrReadOnly
from .models import (
    Diario, Tomo, Capitulo, CapituloImagen, CapituloSlap, CapituloPista,
    Personaje, Conversacion, Mensaje, Expediente, ExpedienteImagen, SaludoEditora,
    Presentacion, Slide, Tablero, RecuerdoLeticia, InstagramPerfil, InstagramPost,
    MercadoUmbralNoticia, MercadoUmbralCompra, MercadoUmbralCyborg, MercadoUmbralHumano,
    MercadoUmbralHumanoImagen, Bienvenida, LibroVisitas, Seguridad, Desktop, Susurro, ClaveAcceso
)
from .serializers import (
    DiarioSerializer, TomoSerializer, CapituloSerializer,
    CapituloImagenSerializer, CapituloSlapSerializer, CapituloPistaSerializer,
    PersonajeSerializer, ConversacionSerializer, MensajeSerializer,
    ExpedienteSerializer, ExpedienteImagenSerializer, SaludoEditoraSerializer,
    PresentacionSerializer, SlideSerializer, TableroSerializer, RecuerdoLeticiaSerializer,
    InstagramPerfilSerializer, InstagramPostSerializer,
    MercadoUmbralNoticiaSerializer, MercadoUmbralCompraSerializer,
    MercadoUmbralCyborgSerializer, MercadoUmbralHumanoSerializer,
    MercadoUmbralHumanoImagenSerializer, BienvenidaSerializer, LibroVisitasSerializer, SeguridadSerializer, DesktopSerializer, SusurroSerializer, ClaveAccesoSerializer
)

class SeguridadViewSet(viewsets.ModelViewSet):
    queryset = Seguridad.objects.all()
    serializer_class = SeguridadSerializer
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['id']

    @action(detail=False, methods=['get'])
    def random_challenge(self, request):
        """Returns a random 'santo' (challenge) without the 'sena'."""
        count = self.queryset.count()
        if count == 0:
            return Response({"error": "No challenges available"}, status=status.HTTP_404_NOT_FOUND)
        
        import random
        random_index = random.randint(0, count - 1)
        # Efficient random pick might be OrderBy('?') but for small tables this is fine
        # Using index slicing if IDs are contiguous, but they might not be.
        # Better: random choice from IDs
        pks = self.queryset.values_list('pk', flat=True)
        random_pk = random.choice(pks)
        obj = self.queryset.get(pk=random_pk)
        
        return Response({
            "id": obj.id,
            "santo": obj.santo
        })

    @action(detail=True, methods=['post'])
    def verify_response(self, request, pk=None):
        """Verifies the 'sena' (response) for a given challenge ID."""
        obj = self.get_object()
        user_response = request.data.get('sena', '').strip()
        
        if not user_response:
             return Response({"error": "Respuesta (sena) requerida"}, status=status.HTTP_400_BAD_REQUEST)

        import unicodedata
        import re

        def normalize_text(text):
            if not text: return ""
            # Lowercase
            text = text.lower()
            # Remove accents
            text = unicodedata.normalize('NFD', text).encode('ascii', 'ignore').decode("utf-8")
            # Remove non-alphanumeric (keep only letters and numbers)
            text = re.sub(r'[^a-z0-9]', '', text)
            return text

        normalized_input = normalize_text(user_response)
        normalized_stored = normalize_text(obj.sena)

        if normalized_input == normalized_stored:
            return Response({"status": "correct", "message": "Acceso concedido"})
        else:
            return Response({"status": "incorrect", "message": "Respuesta incorrecta"}, status=status.HTTP_403_FORBIDDEN)

class DesktopViewSet(viewsets.ModelViewSet):
    queryset = Desktop.objects.all()
    serializer_class = DesktopSerializer
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['nombre', 'id']

class SusurroViewSet(viewsets.ModelViewSet):
    queryset = Susurro.objects.all()
    serializer_class = SusurroSerializer
    ordering_fields = ['seccion', 'id']

    def get_queryset(self):
        queryset = super().get_queryset()
        if not self.request.user.is_staff:
             queryset = queryset.filter(activado=True)
        return queryset

class DiarioViewSet(viewsets.ModelViewSet):
    queryset = Diario.objects.all()
    serializer_class = DiarioSerializer

class TomoViewSet(viewsets.ModelViewSet):
    queryset = Tomo.objects.all()
    serializer_class = TomoSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['diario']

class CapituloViewSet(viewsets.ModelViewSet):
    queryset = Capitulo.objects.all()
    serializer_class = CapituloSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['tomo', 'es_demo', 'is_vip']
    ordering_fields = ['orden', 'id']

    def get_queryset(self):
        queryset = super().get_queryset()
        # En el manager siempre queremos ver todos los capítulos (activos e inactivos)
        # El frontend público (ultragen_site_simple) ya filtra por is_active en su propio código
        return queryset

class CapituloImagenViewSet(viewsets.ModelViewSet):
    queryset = CapituloImagen.objects.all()
    serializer_class = CapituloImagenSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['capitulo']

class CapituloSlapViewSet(viewsets.ModelViewSet):
    queryset = CapituloSlap.objects.all()
    serializer_class = CapituloSlapSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['capitulo']

class CapituloPistaViewSet(viewsets.ModelViewSet):
    queryset = CapituloPista.objects.all()
    serializer_class = CapituloPistaSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['capitulo']

class PersonajeViewSet(viewsets.ModelViewSet):
    queryset = Personaje.objects.all()
    serializer_class = PersonajeSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['diario', 'tomo', 'capitulo']

class ConversacionViewSet(viewsets.ModelViewSet):
    queryset = Conversacion.objects.all()
    serializer_class = ConversacionSerializer

    def create(self, request, *args, **kwargs):
        personajes_ids = request.data.get('personajes', [])
        if len(personajes_ids) != 2:
            from rest_framework.response import Response
            from rest_framework import status
            return Response(
                {"error": "Se requieren exactamente 2 personajes para crear una conversación."},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Create conversation
        conversacion = Conversacion.objects.create()
        conversacion.personajes.set(personajes_ids)
        
        serializer = self.get_serializer(conversacion)
        from rest_framework.response import Response
        from rest_framework import status
        return Response(serializer.data, status=status.HTTP_201_CREATED)

class MensajeViewSet(viewsets.ModelViewSet):
    queryset = Mensaje.objects.all()
    serializer_class = MensajeSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['conversacion', 'personaje']
    ordering_fields = ['orden', 'fecha_simulada', 'hora_simulada']

class ExpedienteViewSet(viewsets.ModelViewSet):
    queryset = Expediente.objects.all()
    serializer_class = ExpedienteSerializer
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['orden', 'id']

class ExpedienteImagenViewSet(viewsets.ModelViewSet):
    queryset = ExpedienteImagen.objects.all()
    serializer_class = ExpedienteImagenSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['expediente']

class SaludoEditoraViewSet(viewsets.ModelViewSet):
    queryset = SaludoEditora.objects.all()
    serializer_class = SaludoEditoraSerializer

class PresentacionViewSet(viewsets.ModelViewSet):
    queryset = Presentacion.objects.all()
    serializer_class = PresentacionSerializer

class SlideViewSet(viewsets.ModelViewSet):
    queryset = Slide.objects.all()
    serializer_class = SlideSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['presentacion']

class TableroViewSet(viewsets.ModelViewSet):
    queryset = Tablero.objects.all()
    serializer_class = TableroSerializer
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['orden', 'id']

class RecuerdoLeticiaViewSet(viewsets.ModelViewSet):
    queryset = RecuerdoLeticia.objects.all()
    serializer_class = RecuerdoLeticiaSerializer
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['orden', 'id']

class InstagramPerfilViewSet(viewsets.ModelViewSet):
    queryset = InstagramPerfil.objects.all()
    serializer_class = InstagramPerfilSerializer

class InstagramPostViewSet(viewsets.ModelViewSet):
    queryset = InstagramPost.objects.all()
    serializer_class = InstagramPostSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['perfil']
    ordering_fields = ['fecha', 'id']

class MercadoUmbralNoticiaViewSet(viewsets.ModelViewSet):
    queryset = MercadoUmbralNoticia.objects.all()
    serializer_class = MercadoUmbralNoticiaSerializer
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['fecha', 'id']

class MercadoUmbralCompraViewSet(viewsets.ModelViewSet):
    queryset = MercadoUmbralCompra.objects.all()
    serializer_class = MercadoUmbralCompraSerializer
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['fecha', 'id']
    
    def get_queryset(self):
        queryset = super().get_queryset()
        if not self.request.user.is_staff:
             queryset = queryset.filter(is_active=True)
        return queryset

class MercadoUmbralCyborgViewSet(viewsets.ModelViewSet):
    queryset = MercadoUmbralCyborg.objects.all()
    serializer_class = MercadoUmbralCyborgSerializer
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['nombre', 'id']

    def get_queryset(self):
        queryset = super().get_queryset()
        if not self.request.user.is_staff:
             queryset = queryset.filter(disponible=True, is_active=True)
        return queryset

class MercadoUmbralHumanoViewSet(viewsets.ModelViewSet):
    queryset = MercadoUmbralHumano.objects.all()
    serializer_class = MercadoUmbralHumanoSerializer
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['nombre', 'id']

    def get_queryset(self):
        queryset = super().get_queryset()
        if not self.request.user.is_staff:
             queryset = queryset.filter(is_active=True)
        return queryset

class MercadoUmbralHumanoImagenViewSet(viewsets.ModelViewSet):
    queryset = MercadoUmbralHumanoImagen.objects.all()
    serializer_class = MercadoUmbralHumanoImagenSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['humano']
    ordering_fields = ['orden', 'id']

class BienvenidaViewSet(viewsets.ModelViewSet):
    queryset = Bienvenida.objects.all()
    serializer_class = BienvenidaSerializer
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['fecha', 'id']

class LibroVisitasViewSet(viewsets.ModelViewSet):
    queryset = LibroVisitas.objects.all()
    serializer_class = LibroVisitasSerializer
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['fecha', 'id']

from django.apps import apps
from django.http import HttpResponse
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework import status
import openpyxl
from openpyxl.utils import get_column_letter

class UniversalLoaderViewSet(viewsets.ViewSet):
    permission_classes = [IsAdminUser]
    """
    A ViewSet for listing models, downloading templates, and importing data.
    """

    def list(self, request):
        """List all models in the 'core' app."""
        app_config = apps.get_app_config('core')
        models = []
        for model in app_config.get_models():
            models.append({
                'name': model.__name__,
                'verbose_name': model._meta.verbose_name.title()
            })
        return Response(models)

    @action(detail=False, methods=['get'], url_path='data/(?P<model_name>[^/.]+)')
    def list_data(self, request, model_name=None):
        """List data for a specific model."""
        try:
            model = apps.get_model('core', model_name)
            # Simple serialization: list of dicts
            # For production, use dynamic serializers or values()
            # data = list(model.objects.all().values())
            data = list(model.objects.all().values())
            return Response(data)
        except LookupError:
            return Response({"error": "Model not found"}, status=status.HTTP_404_NOT_FOUND)

    @action(detail=False, methods=['get'], url_path='template/(?P<model_name>[^/.]+)')
    def get_template(self, request, model_name=None):
        """Download an empty .xlsx template for the model."""
        try:
            model = apps.get_model('core', model_name)
        except LookupError:
            return Response({"error": "Model not found"}, status=status.HTTP_404_NOT_FOUND)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = model_name

        # Get fields, excluding auto-created (like ID) and reverse relations
        fields = [f.name for f in model._meta.get_fields() if not f.auto_created and f.concrete]
        
        # Write header
        for col_num, field_name in enumerate(fields, 1):
            col_letter = get_column_letter(col_num)
            ws[f'{col_letter}1'] = field_name

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={model_name}_template.xlsx'
        wb.save(response)
        return response

    @action(detail=False, methods=['post'], url_path='import/(?P<model_name>[^/.]+)')
    def import_data(self, request, model_name=None):
        """Import data from .xlsx."""
        try:
            model = apps.get_model('core', model_name)
        except LookupError:
            return Response({"error": "Model not found"}, status=status.HTTP_404_NOT_FOUND)

        file = request.FILES.get('file')
        if not file:
            return Response({"error": "No file provided"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            
            headers = [cell.value for cell in ws[1]]
            rows_created = 0
            errors = []

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                if not any(row): continue # Skip empty rows

                row_data = dict(zip(headers, row))
                # Basic handling: ignore fields not in model, handle simple types
                # Complex types (ForeignKey, ImageField) might fail or need ID/path
                
                # Filter out keys that might be None or not in model fields
                model_fields = {f.name for f in model._meta.get_fields()}
                clean_data = {k: v for k, v in row_data.items() if k in model_fields and v is not None}

                try:
                    model.objects.create(**clean_data)
                    rows_created += 1
                except Exception as e:
                    errors.append(f"Row {row_idx}: {str(e)}")

            return Response({
                "status": "success",
                "rows_created": rows_created,
                "errors": errors
            })

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

class DashboardStatsViewSet(viewsets.ViewSet):
    """
    ViewSet to return aggregated stats for the dashboard.
    """
    def list(self, request):
        data = {
            "desktop": {
                "total": Desktop.objects.count()
            },
            "diarios": {
                "total": Diario.objects.count(),
                "tomos": Tomo.objects.count(),
                "capitulos": Capitulo.objects.count(),
                "personajes": Personaje.objects.count()
            },
            "movil": {
                "conversaciones": Conversacion.objects.count(),
                "mensajes": Mensaje.objects.count(),
                "perfiles": InstagramPerfil.objects.count(),
                "posts": InstagramPost.objects.count(),
                "total": Conversacion.objects.count() + InstagramPerfil.objects.count()
            },
            "expedientes": {
                "total": Expediente.objects.count(),
                "imagenes": ExpedienteImagen.objects.count()
            },
            "tablet": {
                "presentaciones": Presentacion.objects.count(),
                "slides": Slide.objects.count(),
                "total": Presentacion.objects.count()
            },
            "pistas": {
                "tableros": Tablero.objects.count(),
                "total": Tablero.objects.count()
            },
            "laptop": {
                "cyborgs": MercadoUmbralCyborg.objects.count(),
                "humanos": MercadoUmbralHumano.objects.count(),
                "compras": MercadoUmbralCompra.objects.count(),
                "noticias": MercadoUmbralNoticia.objects.count(),
                "total": MercadoUmbralCyborg.objects.count() + MercadoUmbralHumano.objects.count() + MercadoUmbralCompra.objects.count() + MercadoUmbralNoticia.objects.count()
            },
            "popup": {
                "saludos": SaludoEditora.objects.count(),
                "recuerdos": RecuerdoLeticia.objects.count(),
                "bienvenidas": Bienvenida.objects.count(),
                "visitas": LibroVisitas.objects.count(),
                "seguridad": Seguridad.objects.count(),
                "total": SaludoEditora.objects.count() + RecuerdoLeticia.objects.count() + Bienvenida.objects.count() + LibroVisitas.objects.count() + Seguridad.objects.count()
            }
        }
        return Response(data)

from django.http import HttpResponse
from django.conf import settings
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAdminUser

@api_view(['GET'])
@permission_classes([IsAdminUser])
def debug_storage_view(request):
    lines = []
    lines.append("--- DIAGNOSTICO DE PRODUCCION ---")
    lines.append(f"DEBUG: {settings.DEBUG}")
    lines.append(f"DEFAULT_FILE_STORAGE: {getattr(settings, 'DEFAULT_FILE_STORAGE', 'No definido')}")
    try:
        lines.append(f"STORAGES: {settings.STORAGES}")
    except AttributeError:
        lines.append("STORAGES: No definido")
    
    lines.append(f"Storage Activo: {default_storage.__class__.__name__}")
    
    # Test upload
    try:
        filename = 'cloud_test_web.txt'
        if default_storage.exists(filename):
            default_storage.delete(filename)
        path = default_storage.save(filename, ContentFile(b'verificacion web'))
        url = default_storage.url(path)
        lines.append(f"Prueba de subida exitosa. URL: {url}")
        
        if 'cloudinary' in url:
            lines.append(">>> RESULTADO: CLOUDINARY ACTIVADO.")
        else:
            lines.append(">>> RESULTADO: ALERTA - ALMACENAMIENTO LOCAL.")
    except Exception as e:
        lines.append(f"Error subida: {str(e)}")
        
    return HttpResponse("\n".join(lines), content_type="text/plain")

class ClaveAccesoViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = ClaveAcceso.objects.all()
    serializer_class = ClaveAccesoSerializer

    @action(detail=False, methods=['post'])
    def validar(self, request):
        clave_input = request.data.get('clave')
        if not clave_input:
            return Response({'error': 'No se proporcionó ninguna clave'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            # Check for any valid key with this code (case-insensitive)
            obj = ClaveAcceso.objects.get(clave__iexact=clave_input)
            
            if obj.esta_valida():
                return Response({
                    'status': 'valida',
                    'tipo': obj.tipo,
                    'nombre': obj.nombre
                })
            else:
                return Response({
                    'status': 'expirada',
                    'mensaje': 'Esta clave ya no está vigente'
                }, status=status.HTTP_403_FORBIDDEN)
                
        except ClaveAcceso.DoesNotExist:
            return Response({
                'status': 'invalida',
                'mensaje': 'Clave incorrecta'
            }, status=status.HTTP_404_NOT_FOUND)
